
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill

file_path = "The_Sports_Memorabilia_Store_Financial_Model.xlsx"

try:
    # 1. Define Start-up Costs (Explicitly as per user request/screenshot)
    # Total must equal 10,000
    startup_data = {
        'Category': ['Capital Expenditure (CapEx)', 'Capital Expenditure (CapEx)', 'Capital Expenditure (CapEx)', 
                     'Pre-Start-up Expenses', 'Pre-Start-up Expenses', 'TOTAL'],
        'Item': ['Initial Inventory', 'B2B sales', 'Packaging', 
                 'Brand Identity & Design', 'Legal & Incorporation Fees', 'Total Start-up Costs'],
        'Cost (£)': [6500, 1500, 1300, 200, 500, 10000],
        'Notes': ['Signing with Chris Eubank, items signed and frames', 'Rhys Barker - B2B pipeline build', 
                  'Packaging includes luxury box, styrofoam and outer shipping box.', 'Logo, assets, guidelines', 
                  'Companies House, Contracts', '']
    }
    df_startup = pd.DataFrame(startup_data)
    print("Defined Start-up Costs (Total £10,000).")

    # Load other sheets to preserve edits
    xls = pd.ExcelFile(file_path)
    # df_startup is now hardcoded, so we don't read it
    df_opex = pd.read_excel(xls, 'Operating Expenditure')
    df_cogs = pd.read_excel(xls, 'Cost of Goods Sold')
    df_non_opex = pd.read_excel(xls, 'Non-Operating Expenditure')
    df_projections = pd.read_excel(xls, 'Projections (5 Years)')

    print("Successfully loaded existing user data (OpEx, COGS, etc).")

    # 2. Define New "Signing Event" Assumptions
    # User requested specific breakdown for travel/hotel per signing.
    signing_cost_data = {
        'Item': ['Travel (Train/Flight/Fuel)', 'Accommodation (Hotel)', 'Subsistence (Meals)', 
                 'Logistics (Secure Van/Courier)', 'Contingency'],
        'Cost Per Trip (£)': [150, 150, 50, 100, 100],
        'Notes': ['Avg cost to reach athlete', '1 night stay', 'Staff meals', 'Transporting bulk items securely', 'Unexpected costs']
    }
    df_signing_costs = pd.DataFrame(signing_cost_data)
    cost_per_signing = df_signing_costs['Cost Per Trip (£)'].sum()
    
    # Add Total row
    df_signing_costs.loc[len(df_signing_costs)] = ['TOTAL PER EVENT', cost_per_signing, '']
    
    print(f"Calculated Cost Per Signing Trip: £{cost_per_signing}")

    # 3. Build Year 1 Monthly Breakdown (Driver-Based)
    # Strategy: 5 Major Signings. Focus is on "selling all products available" immediately.
    # Drivers based on "Chris Eubank" example provided.
    
    # 3. Year 1 Monthly Breakdown - Cash Flow Focused & Realistic Sales Curve
    print("Building Realistic Cash Flow Model (Sales Spread)...")
    
    months = [f'Month {i}' for i in range(1, 13)]
    
    # --- DRIVERS ---
    
    # 1. Revenue with Spread (The "Tail")
    # Drop Schedule: M2, M4, M6, M9, M11
    # 1-indexed months: 2, 4, 6, 9, 11
    drop_months = [2, 4, 6, 9, 11]
    
    target_drop_revenue = 43800
    
    # Sales Curve: 40% (M), 30% (M+1), 20% (M+2), 10% (M+3)
    sales_curve = [0.40, 0.30, 0.20, 0.10]
    
    revenue_monthly = [0.0] * 12
    
    for start_month in drop_months:
        start_idx = start_month - 1 # 0-indexed
        for curve_idx, pct in enumerate(sales_curve):
            current_idx = start_idx + curve_idx
            if current_idx < 12:
                revenue_monthly[current_idx] += target_drop_revenue * pct
                
    # Organic is 0 as requested
    organic_revenue = [0.0] * 12
    
    # 2. Inventory / COGS (Cash Outflow)
    # Drop 1 (M2) inventory is covered by the £10k Startup Costs (M1).
    # Drop 2, 3, 4, 5 Inventory (£11,500) must be purchased upfront (Cash Out).
    # Assumption: Purchase happens in the month of the drop (Just in Time) or month prior?
    # Let's assume Month of Drop for simplicity of the model, or user can adjust.
    
    inventory_cash_out = [0.0] * 12
    cogs_per_drop_cash = 11500
    
    for i, m_num in enumerate(drop_months):
        idx = m_num - 1
        if i == 0:
            # First drop - Covered by Startup Capital
            inventory_cash_out[idx] = 0
        else:
            # Subsequent drops - Cash purchase
            inventory_cash_out[idx] = cogs_per_drop_cash
            
    # 3. OpEx (Cash Outflow)
    # Salaries: 1.5k (M1), 4k (M2+)
    salaries = [1500 if i==0 else 4000 for i in range(12)]
    
    # Ads: 0 (M1), 4k (M2+)
    ads = [0 if i==0 else 4000 for i in range(12)]
    
    # Content: 500 flat
    content = [500] * 12
    
    # Base Fixed (Tech/Storage/Legal) - Calculated from sheet earlier
    # Re-calculate to be safe
    mask_other = ~df_opex['Expense Item'].str.contains('Marketing|Salaries|Founders', case=False, na=False)
    df_opex_other = df_opex[mask_other]
    base_fixed_val = df_opex_other[pd.to_numeric(df_opex_other['Monthly Cost (£)'], errors='coerce').notnull()]['Monthly Cost (£)'].sum()
    fixed_costs = [base_fixed_val] * 12
    
    # Logistics (Travel) - Only on Drop Months
    logistics = [0.0] * 12
    for m in drop_months:
        logistics[m-1] = cost_per_signing
        
    # 4. Cash Flow & Balance
    investment = [10000] + [0]*11
    # Startup Costs are purely the specific items listed (Inventory, Legal etc) = 10k
    # We treat this as a one-off outflow in M1.
    startup_outflow = [10000] + [0]*11
    
    opening_bal = []
    closing_bal = []
    net_flow_monthly = []
    
    curr = 0
    for i in range(12):
        # Total Cash In (Revenue + Investment)
        total_cash_in_month = revenue_monthly[i] + investment[i]
        
        # Total Cash Out (Inventory + OpEx + Startup Outflow)
        total_cash_out_month = inventory_cash_out[i] + salaries[i] + ads[i] + content[i] + fixed_costs[i] + logistics[i] + startup_outflow[i]
        
        # Net Cash Flow
        net = total_cash_in_month - total_cash_out_month
        
        close = curr + net
        
        opening_bal.append(curr)
        closing_bal.append(close)
        net_flow_monthly.append(net)
        
        curr = close

    # --- DATAFRAME CONSTRUCTION ---
    # Clear "Cash In" vs "Cash Out" Layout
    
    data_rows = [
        # CASH IN
        ['CASH IN: Product Sales (Drops)', *revenue_monthly],
        ['CASH IN: Injection/Investment', *investment],
        ['TOTAL CASH IN', *[r+i for r,i in zip(revenue_monthly, investment)]],
        ['', *[''] * 12],
        
        # CASH OUT
        ['CASH OUT: Start-up Costs (One-off)', *startup_outflow],
        ['CASH OUT: Inventory Purchases (Restock)', *inventory_cash_out],
        ['CASH OUT: Staff & Founder Fees', *salaries],
        ['CASH OUT: Paid Advertising', *ads],
        ['CASH OUT: Content Creation', *content],
        ['CASH OUT: Signing Logistics', *logistics],
        ['CASH OUT: Fixed Overheads (Tech/Storage)', *fixed_costs],
        ['TOTAL CASH OUT', *[startup_outflow[i] + inventory_cash_out[i] + salaries[i] + ads[i] + content[i] + fixed_costs[i] + logistics[i] for i in range(12)]],
        ['', *[''] * 12],
        
        # BALANCE
        ['NET CASH FLOW', *net_flow_monthly],
        ['Opening Bank Balance', *opening_bal],
        ['Closing Bank Balance', *closing_bal]
    ]

    cols = ['Metric'] + months
    df_y1_monthly = pd.DataFrame(data_rows, columns=cols)
    
    # Calculate Year 1 Totals
    y1_values = []
    for idx, row in df_y1_monthly.iterrows():
        metric = str(row['Metric'])
        if metric == '' or (isinstance(row[months[0]], str) and row[months[0]] == ''):
             y1_values.append('')
        elif 'Balance' in metric:
             # Closing Balance takes the last month's value
             if 'Closing' in metric:
                 y1_values.append(row[months[-1]])
             elif 'Opening' in metric:
                 y1_values.append(row[months[0]]) # Or just N/A for a total
             else:
                 y1_values.append(sum([float(x) for x in row[months] if isinstance(x, (int, float))]))
        else:
             y1_values.append(sum([float(x) for x in row[months] if isinstance(x, (int, float))]))
             
    df_y1_monthly['YEAR 1 TOTAL'] = y1_values

    # 4. Synchronize Projections (Approximate Mapping)
    # Since we moved to Cash Basis, P&L Sync is slightly different but we map Totals.
    
    total_rev = df_y1_monthly[df_y1_monthly['Metric'] == 'CASH IN: Product Sales (Drops)']['YEAR 1 TOTAL'].values[0]
    total_inv_spend = df_y1_monthly[df_y1_monthly['Metric'] == 'CASH OUT: Inventory Purchases (Restock)']['YEAR 1 TOTAL'].values[0]
    
    # COGS in P&L is typically matched to sales. 
    # For simplicity, we'll map the Cash Inventory Spend + Startup Inventory portion as COGS.
    # Startup Inventory was 6500.
    total_cogs_est = total_inv_spend + 6500 # This assumes the initial inventory is fully sold in Year 1
    
    # Sum all OpEx cash outflows
    total_opex_est = (df_y1_monthly[df_y1_monthly['Metric'] == 'CASH OUT: Staff & Founder Fees']['YEAR 1 TOTAL'].values[0] +
                      df_y1_monthly[df_y1_monthly['Metric'] == 'CASH OUT: Paid Advertising']['YEAR 1 TOTAL'].values[0] +
                      df_y1_monthly[df_y1_monthly['Metric'] == 'CASH OUT: Content Creation']['YEAR 1 TOTAL'].values[0] +
                      df_y1_monthly[df_y1_monthly['Metric'] == 'CASH OUT: Signing Logistics']['YEAR 1 TOTAL'].values[0] +
                      df_y1_monthly[df_y1_monthly['Metric'] == 'CASH OUT: Fixed Overheads (Tech/Storage)']['YEAR 1 TOTAL'].values[0])
    
    # Add pre-startup expenses that are not inventory (Brand Identity 200, Legal 500)
    total_opex_est += 200 + 500

    y1_totals_for_pnl = {
        'TOTAL REVENUE': total_rev,
        'COGS (Inventory & production)': total_cogs_est,
        'GROSS PROFIT': total_rev - total_cogs_est,
        'TOTAL OPERATING EXPENSES': total_opex_est,
        'EBITDA (Operating Profit)': (total_rev - total_cogs_est) - total_opex_est
    }

    # Formula Injection (Vertical)
    from openpyxl.utils import get_column_letter
    
    # Indices changes based on new rows
    # DF Indices (0-based) -> Excel Rows (1-based)
    # CASH IN:
    # Product Sales: DF 0 -> Excel 2
    # Investment: DF 1 -> Excel 3
    # TOTAL CASH IN: DF 2 -> Excel 4
    
    # CASH OUT:
    # Start-up Costs: DF 4 -> Excel 6
    # Inventory: DF 5 -> Excel 7
    # Staff: DF 6 -> Excel 8
    # Ads: DF 7 -> Excel 9
    # Content: DF 8 -> Excel 10
    # Logistics: DF 9 -> Excel 11
    # Fixed: DF 10 -> Excel 12
    # TOTAL CASH OUT: DF 11 -> Excel 13
    
    # BALANCE:
    # NET CASH FLOW: DF 13 -> Excel 15
    # Opening Bal: DF 14 -> Excel 16
    # Closing Bal: DF 15 -> Excel 17
    
    idx_total_cash_in = 2
    idx_total_cash_out = 11
    idx_net_cash_flow = 13
    idx_opening_bal = 14
    idx_closing_bal = 15
    
    for col_idx, col_name in enumerate(months):
        col = get_column_letter(col_idx + 2)
        
        # Total CASH IN
        df_y1_monthly.at[idx_total_cash_in, col_name] = f"=SUM({col}2:{col}3)"
        
        # Total CASH OUT
        df_y1_monthly.at[idx_total_cash_out, col_name] = f"=SUM({col}6:{col}12)"
        
        # NET CASH FLOW (= In - Out)
        df_y1_monthly.at[idx_net_cash_flow, col_name] = f"={col}4-{col}13"
        
        # BALANCE FORMULAS
        # Opening Balance
        if col_idx == 0:
            # Month 1 Opening is 0
            df_y1_monthly.at[idx_opening_bal, col_name] = 0
        else:
            # Previous Month Closing
            prev_col = get_column_letter(col_idx + 1)
            # Link to Previous Closing (Excel Row 17)
            df_y1_monthly.at[idx_opening_bal, col_name] = f"={prev_col}17"
            
        # Closing Balance (= Opening (16) + Net (15))
        df_y1_monthly.at[idx_closing_bal, col_name] = f"={col}16+{col}15"
        
    print("Injected Vertical Excel Formulas.")
    
    def update_projection(metric_name, value):
        mask = df_projections['Metric'].str.contains(metric_name, case=False, na=False)
        if mask.any():
            df_projections.loc[mask, 'Year 1'] = value
    
    # Use distinct keys from the calc dict for Projections
    # Note: earlier we calculated y1_totals_for_pnl, use that variable. NOT y1_totals.
    update_projection('Revenue', y1_totals_for_pnl.get('TOTAL REVENUE', 0))
    update_projection('COGS', y1_totals_for_pnl.get('COGS (Inventory & production)', 0))
    update_projection('Gross Profit', y1_totals_for_pnl.get('GROSS PROFIT', 0))
    
    # Total OpEx
    update_projection('Operating Expenses', y1_totals_for_pnl.get('TOTAL OPERATING EXPENSES', 0))
    
    update_projection('EBITDA', y1_totals_for_pnl.get('EBITDA (Operating Profit)', 0))
    # Net Profit same as EBITDA for now
    update_projection('Net Profit', y1_totals_for_pnl.get('EBITDA (Operating Profit)', 0)) 

    print("Synchronized Year 1 Projections (Cash Basis adjusted).")

    # 4b. Convert Totals to Excel Formulas
    # User requested actual formulas (e.g. =SUM(...)) in the output file instead of hardcoded numbers.
    # We apply these mutations to the DataFrames *after* using the values for synchronization.
    
    # 1. Start-up Costs Formulas
    df_startup.iloc[-1, df_startup.columns.get_loc('Cost (£)')] = "=SUM(C2:C6)"
    
    # 2. Signing Trip Costs Formulas
    df_signing_costs.iloc[-1, df_signing_costs.columns.get_loc('Cost Per Trip (£)')] = "=SUM(B2:B6)"
    
    # 3. Year 1 Monthly Formulas (Horizontal 'Year Total')
    # We replace the 'YEAR 1 TOTAL' column values with formulas.
    # Columns B to M are Months 1-12. 'YEAR 1 TOTAL' is Column N.
    # Rows start at index 0 -> Excel Row 2.
    
    formula_list_horiz = []
    for idx, row in df_y1_monthly.iterrows():
        excel_row = idx + 2
        metric = row['Metric']
        
        # Handle Spacer Rows
        if not isinstance(metric, str) or metric == '':
            formula_list_horiz.append("")
            continue
        
        if 'Balance' in metric:
            # Closing Balance logic: Just take the Month 12 value
            formula_list_horiz.append(f"=M{excel_row}")
        else:
            # Sum logic for Revenue, Expenses etc
            formula_list_horiz.append(f"=SUM(B{excel_row}:M{excel_row})")
            
    df_y1_monthly['YEAR 1 TOTAL'] = formula_list_horiz
    
    print("Converted static totals to Excel formulas.")

    # 5. Write to Excel (Preserving formatting is hard with pandas, but we will make it clean)
    # Saving to a new file to avoid PermissionError if the user has the original open
    output_path = "The_Sports_Memorabilia_Store_Financial_Model_UPDATED_v2.xlsx"
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_startup.to_excel(writer, sheet_name='Start-up Costs', index=False)
        df_opex.to_excel(writer, sheet_name='Operating Expenditure', index=False)
        df_cogs.to_excel(writer, sheet_name='Cost of Goods Sold', index=False)
        df_signing_costs.to_excel(writer, sheet_name='Signing Trip Costs', index=False) # NEW TAB
        df_y1_monthly.to_excel(writer, sheet_name='Year 1 Monthly (Scalability)', index=False) # NEW TAB
        df_projections.to_excel(writer, sheet_name='Projections (5 Years)', index=False)
        df_non_opex.to_excel(writer, sheet_name='Non-Operating Expenditure', index=False)

        # Basic Formatting
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            # Auto-adjust column widths roughly
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

    print(f"Financial Model updated successfully at: {os.path.abspath(output_path)}")

except Exception as e:
    print(f"Error updating financial model: {e}")
